import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ==============================
# LIST LINK SCRAPING
# ==============================
links = [
    "https://www.cnbcindonesia.com/entrepreneur/20251122115747-25-687621/semua-rata-tanah-warga-lihat-banjir-lahar-semeru-bak-tsunami-raksasa",
    "https://www.cnbcindonesia.com/market/20230208142031-17-412137/bos-the-fed-sebut-terjadi-disinflasi-di-as-maksudnya-apa-ya",
    "https://www.cnbcindonesia.com/lifestyle/20230308164031-33-420030/10-hewan-berumur-paling-panjang-di-bumi-ada-yang-hidup-abadi",
    "https://www.cnbcindonesia.com/news/20230504134127-4-434453/progres-pabrik-bauksit-berantakan-pemerintah-wajib-audit",
    "https://www.cnbcindonesia.com/news/20231017130313-4-481231/50-kontrak-blok-migas-dikembalikan-ke-negara-ini-daftarnya"
    # "6",
    # "7",
    # "8",
    # "9",
    # "10",
    # "11",
    # "12",
    # "13",
    # "14",
    # "15",
    # "16",
    # "17",
    # "18",
    # "19",
    # "20",
    # "21",
    # "22",
    # "23",
    # "24",
    # "25",
    # "26",
    # "27",
    # "28",
    # "29",
    # "30",
    # "31",
    # "32",
    # "33",
    # "34",
    # "35",
    # "36",
    # "37",
    # "38",
    # "39",
    # "40",
    # "41",
    # "42",
    # "43",
    # "44",
    # "45",
    # "46",
    # "47",
    # "48",
    # "49",
    # "50",
    # "51",
    # "52",
    # "53",
    # "54",
    # "55",
    # "56",
    # "57",
    # "58",
    # "59",
    # "60",
    # "61",
    # "62",
    # "63",
    # "64",
    # "65",
    # "66",
    # "67",
    # "68",
    # "69",
    # "70",
    # "71",
    # "72",
    # "73",
    # "74",
    # "75",
    # "76",
    # "77",
    # "78",
    # "79",
    # "80",
    # "81",
    # "82",
    # "83",
    # "84",
    # "85",
    # "86",
    # "87",
    # "88",
    # "89",
    # "90",
    # "91",
    # "92",
    # "93",
    # "94",
    # "95",
    # "96",
    # "97",
    # "98",
    # "99",
    # "100",
    # "101",
    # "102",
    # "103",
    # "104",
    # "105",
    # "106",
    # "107",
    # "108",
    # "109",
    # "110",
    # "111",
    # "112",
    # "113",
    # "114",
    # "115",
    # "116",
    # "117",
    # "118",
    # "119",
    # "120",
]

hasil = []

# ==============================
# SCRAPING DATA
# ==============================
for i, url in enumerate(links, start=1):
    try:
        print(f"Memproses {i}/{len(links)}: {url}")

        headers = {"User-Agent": "Mozilla/5.0"}
        page = requests.get(url, headers=headers, timeout=10)
        soup = BeautifulSoup(page.text, "html.parser")

        judul = soup.find("h1").text.strip() if soup.find("h1") else "-"

        penulis_div = soup.find("div", class_="mb-1 text-base font-semibold")
        penulis = penulis_div.get_text(strip=True).replace("CNBC Indonesia", "").replace(",", "").strip() if penulis_div else "-"

        tanggal_div = soup.find("div", class_="text-cm text-gray")
        tanggal = tanggal_div.get_text(strip=True) if tanggal_div else "-"

        isi = " ".join([p.text.strip() for p in soup.find_all("p")])

        hasil.append({
            "No": i,
            "Judul Berita": judul,
            "Isi Lengkap Berita": isi,
            "Tanggal Publikasi": tanggal,
            "Nama Penulis": penulis,
            "Link": url
        })

        time.sleep(1)

    except Exception as e:
        print("Error:", e)

# ==============================
# SIMPAN EXCEL
# ==============================
file_path = "C_Kelompok 7.xlsx"
df = pd.DataFrame(hasil)
df.to_excel(file_path, index=False)

wb = load_workbook(file_path)
ws = wb.active
ws.title = "Hasil Scraping"

# ==============================
# FORMAT HEADER
# ==============================
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# ==============================
# AUTO-FIT & ALIGNMENT
# ==============================
for col in ws.columns:
    col_letter = col[0].column_letter
    header = col[0].value

    if header == "Isi Lengkap Berita":
        ws.column_dimensions[col_letter].width = 50
        for cell in col:
            cell.alignment = Alignment(wrap_text=False, vertical="top")
            cell.border = thin_border
    else:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[col_letter].width = max_length + 5
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border

# ==============================
# SHEET KELOMPOK 7
# ==============================
sheet7 = wb.create_sheet("Kelompok 7")
header_row = ["No", "Nama", "NIM"]
sheet7.append(header_row)

anggota = [
    [1, "Muhammad Alief Adhitya Pratama", "L200220281"],
    [2, "Satria Bintang Bagus Pratama", "L200220284"],
    [3, "Yafi Ariella Widyatama", "L200230003"],
    [4, "x", "L2002x"],
    [5, "x", "L2002x"],
    [6, "x", "L2002x"]
]

for row in anggota:
    sheet7.append(row)

# Format sheet kelompok 7
for cell in sheet7[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

for col in sheet7.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    sheet7.column_dimensions[col[0].column_letter].width = max_length + 5
    for cell in col:
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin_border

wb.save(file_path)
print("Selesai! File tersimpan sebagai:", file_path)
